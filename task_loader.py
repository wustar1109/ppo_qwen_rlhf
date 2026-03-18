"""
Task loader for prompt queue from Excel/CSV.
"""

from __future__ import annotations

import csv
import os
import re
from typing import Any, Dict, List, Optional


def _normalize_key(name: Any) -> str:
    text = str(name or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _canonical_key(name: Any) -> str:
    key = _normalize_key(name)
    aliases = {
        "promptid": "prompt_id",
        "ipromptid": "prompt_id",
        "prompt": "prompt",
        "targetscore": "target_score",
        "targetscores": "target_score",
        "maxretry": "max_retry",
        "maxrety": "max_retry",
        "maxretrys": "max_retry",
    }
    return aliases.get(key, key)


def _to_int(value: Any, default: Optional[int] = None) -> Optional[int]:
    if value is None:
        return default
    text = str(value).strip()
    if text == "":
        return default
    try:
        return int(float(text))
    except Exception:
        return default


def _to_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default
    text = str(value).strip()
    if text == "":
        return default
    try:
        return float(text)
    except Exception:
        return default


def _canonical_row(raw_row: Dict[str, Any]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for k, v in (raw_row or {}).items():
        ckey = _canonical_key(k)
        if not ckey:
            continue
        row[ckey] = v
    return row


def _build_task_from_row(
    row: Dict[str, Any],
    row_index: int,
    default_max_retry: int,
    default_target_score: Optional[float] = None,
) -> Optional[Dict[str, Any]]:
    prompt = str(row.get("prompt", "") or "").strip()
    if not prompt:
        return None

    prompt_id = str(row.get("prompt_id", "") or "").strip()
    if not prompt_id:
        prompt_id = f"row_{row_index}"

    task = {
        "prompt_id": prompt_id,
        "prompt": prompt,
        "target_score": _to_float(row.get("target_score"), default_target_score),
        "max_retry": _to_int(row.get("max_retry"), default_max_retry),
        # keep optional fields for compatibility with existing downstream logging
        "negative_prompt": str(row.get("negative_prompt", "") or "").strip(),
        "category": str(row.get("category", "") or "").strip(),
        "notes": str(row.get("notes", "") or "").strip(),
        "row_index": row_index,
    }
    if task["max_retry"] is None or task["max_retry"] < 1:
        task["max_retry"] = default_max_retry
    return task


def _within_data_row_range(data_row_index: int, start_row: int, end_row: int) -> bool:
    if data_row_index < start_row:
        return False
    if end_row > 0 and data_row_index > end_row:
        return False
    return True


def _load_csv_tasks(
    path: str,
    start_row: int,
    end_row: int,
    default_max_retry: int,
    default_target_score: Optional[float],
) -> List[Dict[str, Any]]:
    tasks: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for i, raw_row in enumerate(reader, start=1):
            if not _within_data_row_range(i, start_row, end_row):
                continue
            row = _canonical_row(raw_row)
            task = _build_task_from_row(
                row=row,
                row_index=i,
                default_max_retry=default_max_retry,
                default_target_score=default_target_score,
            )
            if task is not None:
                tasks.append(task)
    return tasks


def _load_xlsx_tasks(
    path: str,
    sheet_name: Optional[str],
    start_row: int,
    end_row: int,
    default_max_retry: int,
    default_target_score: Optional[float],
) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(f"Sheet '{sheet_name}' not found in {path}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return []
        header_keys = [_canonical_key(h) for h in headers]

        tasks: List[Dict[str, Any]] = []
        for i, values in enumerate(rows, start=1):
            if not _within_data_row_range(i, start_row, end_row):
                continue
            row_map: Dict[str, Any] = {}
            for idx, key in enumerate(header_keys):
                if not key:
                    continue
                value = values[idx] if idx < len(values) else None
                row_map[key] = value

            task = _build_task_from_row(
                row=row_map,
                row_index=i,
                default_max_retry=default_max_retry,
                default_target_score=default_target_score,
            )
            if task is not None:
                tasks.append(task)
        return tasks
    finally:
        workbook.close()


def load_prompt_tasks(
    task_path: str,
    sheet_name: Optional[str] = None,
    start_row: int = 1,
    end_row: int = 0,
    default_max_retry: int = 3,
    default_target_score: Optional[float] = None,
) -> List[Dict[str, Any]]:
    """
    Load prompt tasks from .xlsx or .csv.

    Required columns: prompt_id, prompt.
    Optional columns: target_score, max_retry.
    Also accepts aliases: targetscore, maxrety.
    """
    if not task_path:
        raise RuntimeError("task_path is required")
    if not os.path.exists(task_path):
        raise RuntimeError(f"Task file not found: {task_path}")

    start_row = max(1, int(start_row or 1))
    end_row = int(end_row or 0)
    default_max_retry = max(1, int(default_max_retry or 3))

    ext = os.path.splitext(task_path)[1].lower()
    if ext == ".csv":
        tasks = _load_csv_tasks(
            path=task_path,
            start_row=start_row,
            end_row=end_row,
            default_max_retry=default_max_retry,
            default_target_score=default_target_score,
        )
    elif ext in (".xlsx", ".xlsm"):
        tasks = _load_xlsx_tasks(
            path=task_path,
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            default_max_retry=default_max_retry,
            default_target_score=default_target_score,
        )
    else:
        raise RuntimeError(f"Unsupported task file format: {ext}. Use .xlsx or .csv")

    return tasks
